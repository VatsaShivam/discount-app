#!/usr/bin/env python3
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path

OUT = Path('versions.xlsx')

# args: version, status, deploy_url, commit_sha
version = sys.argv[1] if len(sys.argv) > 1 else 'unknown'
status = sys.argv[2] if len(sys.argv) > 2 else 'deployed'
deploy_url = sys.argv[3] if len(sys.argv) > 3 else ''
commit_sha = sys.argv[4] if len(sys.argv) > 4 else ''

if OUT.exists():
    wb = load_workbook(OUT)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(['timestamp', 'version', 'status', 'deploy_url', 'commit_sha'])

ws.append([datetime.utcnow().isoformat(), version, status, deploy_url, commit_sha])
wb.save(OUT)
print('Updated', OUT)
